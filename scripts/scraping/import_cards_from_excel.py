import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import openpyxl
import os

def import_cards_from_excel():
    excel_file_path = "cards_cashback_data.xlsx"
    if not os.path.exists(excel_file_path):
        print(f"LỖI: Không tìm thấy file {excel_file_path}. Vui lòng kiểm tra lại!")
        return

    print("Đang kết nối Firestore...")
    key_path = "scripts/scraping/serviceAccountKey.json"
    
    if not os.path.exists(key_path):
        key_path = "serviceAccountKey.json"
        if not os.path.exists(key_path):
            print(f"LỖI: Không tìm thấy file {key_path}. Hãy tải từ Firebase Console!")
            return

    try:
        cred = credentials.Certificate(key_path)
        firebase_admin.initialize_app(cred)
    except ValueError:
        pass
        
    db = firestore.client()
    cards_ref = db.collection("cards")

    print(f"Đang đọc dữ liệu từ {excel_file_path} và cập nhật lên Firebase...")
    
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb.active
    
    updated_count = 0
    
    def parse_float(val):
        if val is None: return None
        if isinstance(val, (int, float)): return float(val)
        val_str = str(val).replace('%', '').strip()
        if not val_str: return None
        try: return float(val_str)
        except ValueError: return val_str
            
    def parse_int(val):
        if val is None: return None
        if isinstance(val, int): return val
        if isinstance(val, float): return int(val)
        val_str = str(val).replace(',', '').replace('.', '').strip()
        if not val_str: return None
        try: return int(val_str)
        except ValueError: return val_str

    for row_idx in range(2, ws.max_row + 1):
        card_id = ws.cell(row=row_idx, column=1).value
        if not card_id: continue
        
        card_name = ws.cell(row=row_idx, column=3).value
        bank_name = ws.cell(row=row_idx, column=2).value
        
        update_data = {}
        
        val_6 = ws.cell(row=row_idx, column=6).value
        if val_6 is not None and str(val_6).strip(): update_data['medicalCashbackRate'] = parse_float(val_6)
        
        val_7 = ws.cell(row=row_idx, column=7).value
        if val_7 is not None and str(val_7).strip(): update_data['gymCashbackRate'] = parse_float(val_7)
        
        val_8 = ws.cell(row=row_idx, column=8).value
        if val_8 is not None and str(val_8).strip(): update_data['shoppingCashbackRate'] = parse_float(val_8)
        
        val_9 = ws.cell(row=row_idx, column=9).value
        if val_9 is not None and str(val_9).strip(): update_data['supermarketCashbackRate'] = parse_float(val_9)
        
        val_10 = ws.cell(row=row_idx, column=10).value
        if val_10 is not None and str(val_10).strip(): update_data['onlineCashbackRate'] = parse_float(val_10)
        
        val_11 = ws.cell(row=row_idx, column=11).value
        if val_11 is not None and str(val_11).strip(): update_data['travelCashbackRate'] = parse_float(val_11)
        
        val_12 = ws.cell(row=row_idx, column=12).value
        if val_12 is not None and str(val_12).strip(): update_data['diningCashbackRate'] = parse_float(val_12)
        
        val_13 = ws.cell(row=row_idx, column=13).value
        if val_13 is not None and str(val_13).strip(): update_data['transportCashbackRate'] = parse_float(val_13)
        
        val_14 = ws.cell(row=row_idx, column=14).value
        if val_14 is not None and str(val_14).strip(): update_data['educationCashbackRate'] = parse_float(val_14)
        
        val_15 = ws.cell(row=row_idx, column=15).value
        if val_15 is not None and str(val_15).strip(): update_data['utilitiesCashbackRate'] = parse_float(val_15)
        
        val_16 = ws.cell(row=row_idx, column=16).value
        if val_16 is not None and str(val_16).strip(): update_data['entertainmentCashbackRate'] = parse_float(val_16)
        
        val_17 = ws.cell(row=row_idx, column=17).value
        if val_17 is not None and str(val_17).strip(): update_data['insuranceCashbackRate'] = parse_float(val_17)
        
        val_18 = ws.cell(row=row_idx, column=18).value
        if val_18 is not None and str(val_18).strip(): update_data['airportLounge'] = str(val_18)
        
        val_19 = ws.cell(row=row_idx, column=19).value
        if val_19 is not None and str(val_19).strip(): update_data['otherCashbackRate'] = parse_float(val_19)
        
        val_20 = ws.cell(row=row_idx, column=20).value
        if val_20 is not None and str(val_20).strip(): update_data['maxCashbackPerMonth'] = parse_int(val_20)
        
        val_21 = ws.cell(row=row_idx, column=21).value
        if val_21 is not None and str(val_21).strip(): update_data['cashbackNote'] = str(val_21)
        
        if update_data:
            try:
                cards_ref.document(str(card_id)).set(update_data, merge=True)
                print(f"Đã cập nhật: {card_name} ({bank_name})")
                updated_count += 1
            except Exception as e:
                print(f"Lỗi khi cập nhật {card_id}: {e}")

    print(f"\n--- THÀNH CÔNG ---")
    print(f"Đã cập nhật thành công {updated_count} thẻ lên Firestore.")

if __name__ == "__main__":
    import_cards_from_excel()
