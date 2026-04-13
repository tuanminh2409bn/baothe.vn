import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def export_cards_to_excel():
    print("Đang kết nối Firestore...")
    key_path = "scripts/scraping/serviceAccountKey.json"
    
    if not os.path.exists(key_path):
        key_path = "serviceAccountKey.json"
        if not os.path.exists(key_path):
            print(f"LỖI: Không tìm thấy file {key_path}. Hãy tải từ Firebase Console!")
            return

    try:
        cred = credentials.Certificate(key_path)
        firebase_admin.initialize_app(cred)
    except ValueError:
        pass
        
    db = firestore.client()

    print("Đang tải dữ liệu thẻ từ Firebase...")
    cards_ref = db.collection("cards")
    docs = cards_ref.stream()

    excel_file_path = "cards_cashback_data.xlsx"
    
    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data_HoanTien"
    
    # Định nghĩa các cột
    headers = [
        "ID Thẻ (KHÔNG SỬA)", 
        "Ngân hàng", 
        "Tên thẻ", 
        "Link Đăng ký / Chi tiết", 
        "Quyền lợi nổi bật (Tham khảo)", 
        "Hoàn tiền Y tế (%)", 
        "Hoàn tiền Gym/Thể thao (%)", 
        "Hoàn tiền Mua sắm (%)", 
        "Hoàn tiền Siêu thị (%)", 
        "Hoàn tiền Mua sắm Online/Shopee/Tiki (%)",
        "Hoàn tiền Du lịch/Khách sạn/Máy bay (%)",
        "Hoàn tiền Ăn uống/Ẩm thực (%)",
        "Hoàn tiền Di chuyển/Grab/Be/Xanh SM (%)",
        "Hoàn tiền Giáo dục (%)",
        "Hoàn tiền Điện/Nước/Hóa đơn (%)",
        "Hoàn tiền Giải trí (%)",
        "Hoàn tiền Bảo hiểm (%)",
        "Phòng chờ sân bay (Có/Không)",
        "Hoàn tiền Chi tiêu (%)",
        "Max Hoàn tiền / Tháng (VNĐ)",
        "Ghi chú thêm"
    ]

    # Style header
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    header_font = Font(bold=True)
    
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Đặt chiều rộng cho cột
    ws.column_dimensions['A'].width = 25  # ID
    ws.column_dimensions['B'].width = 15  # Ngân hàng
    ws.column_dimensions['C'].width = 30  # Tên thẻ
    ws.column_dimensions['D'].width = 40  # Link
    ws.column_dimensions['E'].width = 50  # Quyền lợi
    
    # Các cột % hoàn tiền (F đến S)
    for col_idx in range(6, 20):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 15
        
    ws.column_dimensions[get_column_letter(20)].width = 20  # Max hoàn tiền
    ws.column_dimensions[get_column_letter(21)].width = 40  # Ghi chú

    total_cards = 0
    row_num = 2
    for doc in docs:
        card = doc.to_dict()
        card_id = doc.id
        
        bank_name = card.get("bankName", "")
        name = card.get("name", "")
        apply_url = card.get("applyUrl", "")
        cashback_highlight = card.get("cashbackHighlight", "")
        
        medical = card.get("medicalCashbackRate", "")
        gym = card.get("gymCashbackRate", "")
        shopping = card.get("shoppingCashbackRate", "")
        supermarket = card.get("supermarketCashbackRate", "")
        online = card.get("onlineCashbackRate", "")
        travel = card.get("travelCashbackRate", "")
        dining = card.get("diningCashbackRate", "")
        transport = card.get("transportCashbackRate", "")
        education = card.get("educationCashbackRate", "")
        utilities = card.get("utilitiesCashbackRate", "")
        entertainment = card.get("entertainmentCashbackRate", "")
        insurance = card.get("insuranceCashbackRate", "")
        lounge = card.get("airportLounge", "")
        other = card.get("otherCashbackRate", "")
        max_cashback = card.get("maxCashbackPerMonth", "")
        note = card.get("cashbackNote", "")

        row = [
            card_id, bank_name, name, apply_url, cashback_highlight,
            medical, gym, shopping, supermarket, online, travel, 
            dining, transport, education, utilities, entertainment, 
            insurance, lounge, other, max_cashback, note
        ]
        
        ws.append(row)
        
        for col_idx in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            if col_idx == 5 or col_idx == 21:  # Quyền lợi hoặc Ghi chú
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx >= 6 and col_idx <= 20:  # Các cột % và số tiền
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

        row_num += 1
        total_cards += 1

    wb.save(excel_file_path)

    print(f"\n--- THÀNH CÔNG ---")
    print(f"Đã xuất {total_cards} thẻ ra file chuẩn: {excel_file_path}")

if __name__ == "__main__":
    export_cards_to_excel()
